#!/usr/bin/env python3
# -*- coding: utf-8 -*-

""" Report module. """

import time
import shutil
import logging
import os
import pandas as pandas
from openpyxl import load_workbook
from source.settings import (
    REPORT_TEMPLATE,
    REPORT_DIR
)


def generate_report(apk_filename, app_basic_info, uses_sdk, uses_feature, permission,
                    uses_permission, uses_permission_sdk23, services, receivers, providers):
    """
    Generate the report.

    @param apk_filename: Basic information of an apk file
    @type  apk_filename: list

    @param app_basic_info: Basic information of an apk file
    @type  app_basic_info: list

    @param uses_sdk: uses-sdk of an apk file
    @type  uses_sdk: list

    @param uses_feature: uses-feature of an apk file
    @type  uses_feature: list

    @param permission: permission of an sdk file
    @type  permission: list

    @param uses_permission: uses-permission of an apk file
    @type  uses_permission: list

    @param uses_permission_sdk23: uses-permission-sdk23 of apk file
    @type  uses_permission_sdk23: list

    @param services: services of an apk file
    @type  services: list

    @param receivers: receivers of an apk file
    @type  receivers: list

    @param providers: providers of an apk file
    @type  providers: list
    """
    logging.info('Generating report ...')

    # report filename
    filename = os.path.basename(apk_filename)
    report_filename = os.path.splitext(filename)[0] + '-' + time.strftime("%Y%m%d-%H%M%S") + '.xlsx'

    # copy template report
    report_path = os.path.join(REPORT_DIR, report_filename)
    try:
        shutil.copy(REPORT_TEMPLATE, report_path)
    except OSError as err:
        logging.exception("Failed to copy the file %s - %s", report_filename, err.strerror)

    # write APK Basic Information sheet
    data_frame = pandas.DataFrame(app_basic_info, columns=['package',
                                                           'versionCode',
                                                           'versionName'])
    append_data_sheet(report_path, data_frame, sheet_name='APK Basic Information', header=0,
                      index=False)

    # write <uses-sdk> sheet
    data_frame = pandas.DataFrame(uses_sdk, columns=['minSdkVersion',
                                                     'targetSdkVersion',
                                                     'maxSdkVersion'])
    append_data_sheet(report_path, data_frame, sheet_name='<uses-sdk>', header=0, index=False)

    # write <uses-feature> sheet
    data_frame = pandas.DataFrame(uses_feature, columns=['name',
                                                         'required',
                                                         'glEsVersion'])
    append_data_sheet(report_path, data_frame, sheet_name='<uses-feature>', header=0, index=False)

    # write <permission> sheet
    data_frame = pandas.DataFrame(permission, columns=['name',
                                                       'description',
                                                       'permissionGroup',
                                                       'protectionLevel'])
    append_data_sheet(report_path, data_frame, sheet_name='<permission>', header=0, index=False)

    # write <uses-permission> sheet
    data_frame = pandas.DataFrame(uses_permission, columns=['name',
                                                            'maxSdkVersion'])
    append_data_sheet(report_path, data_frame, sheet_name='<uses-permission>', header=0,
                      index=False)

    # write <uses-permission-sdk23> sheet
    data_frame = pandas.DataFrame(uses_permission_sdk23, columns=['name',
                                                                  'maxSdkVersion'])
    append_data_sheet(report_path, data_frame, sheet_name='<uses-permission>', header=0,
                      index=False)

    # write <service> sheet
    data_frame = pandas.DataFrame(services, columns=['name',
                                                     'description',
                                                     'directBootAware',
                                                     'enabled',
                                                     'exported',
                                                     'foregroundServiceType',
                                                     'isolatedProcess',
                                                     'permission',
                                                     'process'])
    append_data_sheet(report_path, data_frame, sheet_name='<service>', header=0, index=False)

    # write <receiver> sheet
    data_frame = pandas.DataFrame(receivers, columns=['name',
                                                      'directBootAware',
                                                      'enabled',
                                                      'exported',
                                                      'permission',
                                                      'process'])
    append_data_sheet(report_path, data_frame, sheet_name='<receiver>', header=0, index=False)

    # write <provider> sheet
    data_frame = pandas.DataFrame(providers, columns=['name',
                                                      'authorities',
                                                      'directBootAware',
                                                      'enabled',
                                                      'exported',
                                                      'grantUriPermissions',
                                                      'initOrder',
                                                      'multiprocess',
                                                      'permission',
                                                      'process',
                                                      'readPermission',
                                                      'syncable',
                                                      'writePermission'])
    append_data_sheet(report_path, data_frame, sheet_name='<provider>', header=0, index=False)

    logging.info('Generated report file %s', 'report/' + report_filename)


def append_data_sheet(filename, dataframe, sheet_name='New Sheet', startrow=None,
                      truncate_sheet=False, **to_excel_kwargs):
    """
    Append a dataframe to an existing Excel file

    @param filename: Excel path
    @type  filename: str

    @param sheet_name: Sheet name
    @type  sheet_name: str

    @param startrow: Start row
    @type  startrow: str

    @param truncate_sheet: Truncate sheet
    @type  truncate_sheet: bool

    @param to_excel_kwargs: dataframe.to_excel() arguments
    @type  to_excel_kwargs: list
    """

    if os.path.isfile(filename):
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pandas.ExcelWriter(
            filename, engine='openpyxl')  # pylint: disable=abstract-class-instantiated

        try:
            # read the report template
            writer.book = load_workbook(filename)

            # get the last row of the template
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate the sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])
                writer.book.create_sheet(sheet_name, idx)

            # copy sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except OSError as err:
            logging.exception("Failed to add data in the file %s | sheet name: %s | error: %s",
                              filename, sheet_name, err.errno)

        if startrow is None:
            startrow = 0

        # dataframe to excel
        dataframe.to_excel(writer, sheet_name,
                           startrow=startrow, **to_excel_kwargs)

        # save excel
        writer.save()

    else:
        logging.exception('Could not find the report file %s', filename)
